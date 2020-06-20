#*****************************#
#   Author description        #
# Name : Naveen Kumar Xavier  #
# Started : March 15 2020     #
# Version : 0.1               #
#   End of description        #
#*****************************#

import sqlite3
from collections import OrderedDict
import openpyxl


def databasesearch(**kwargs):
    print("kwargs items ", kwargs.items())
    conn = sqlite3.connect(r"/root/") #establishes connection with DB
    print("success")
    #cursor = conn.execute("SELECT * FROM gdb_info WHERE param='%s' AND value = '%s'" % (value1, value2)) --a static query-parameters passed are predefined
    colomnname = conn.execute("select * from table")
    tablehead = list(map(lambda x :x[0], colomnname.description ))
    print(tablehead) #To get table header names
    where_clause = 'WHERE ' + ' AND '.join(['' + k + ' = ?' for k in kwargs.keys()])
    sql = 'SELECT * FROM gdb_info ' + where_clause
    print(sql) #Query statement to be passed
    values = list(kwargs.values())
    print(values)
    cursor = conn.execute(sql,values)
    result = cursor.fetchall()
    print(result)
    conn.close()
    excelsheet(tablehead, result)

def excelsheet(title, data):
    print("excesheet",data)
    print(type(data))
    wb = openpyxl.load_workbook(r"/root/Output.xlsx")
    sheet = wb.active
    colomn=1
    for i in range(len(title)): #this prints table header data to the excel sheet
        sheet.cell(row=1,column=colomn).value = str(title[i])
        colomn+=1
    row=2
    for i in range(len(data)): #this prints queried data to excel sheet
        colomn = 1
        for j in range(len(data[i])):
            sheet.cell(row=row, column= colomn).value = str(data[i][j])
            j+=1
            colomn+=1
        i+=1
        row+=1
    wb.save(filename='Output.xlsx')

def main():
    inp_vals = [('parameter1', None), ('parameter2', None)] #TEST VECTORS
    passeddict = {}
    inp_vals = OrderedDict(inp_vals) # To maintain the input order of dictionary values
    print(inp_vals)
#Getting input from User
    default = None
    for keys in inp_vals:
        inp_vals[keys] = input()
        if inp_vals[keys] == "":
            inp_vals[keys] = default
        #Printing the same
    for keys, values in inp_vals.items():
        print(keys, ":", values)
    for keys, values in inp_vals.items():
        if inp_vals[keys] != None:
            passeddict[keys] = values
    print("passedDict :", passeddict)
    databasesearch(**passeddict)

if __name__ == '__main__':
    main()
